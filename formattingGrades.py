# Team 10 Section 003
# Make a program that will automatically format and summarize the important 
# information about each of the classes they teach from the imported Excel files.

# Import libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import statistics

# Load the Excel workbook
myWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')

# Get the active sheet
currSheet = myWorkbook.active

# Create a new workbook for the formatted data
formattedWorkbook = Workbook()

# Remove the default sheet created in the workbook
formattedWorkbook.remove(formattedWorkbook['Sheet'])

# Identify different classes and create worksheets for classes
class_names = set()

# Find unique class names
for row in currSheet.iter_rows(min_row = 2, values_only=True) :
    # Class name is in first column
    class_name = row[0]
    class_names.add(class_name)

#manking a class names list
classNamesList = []
# Create a worksheet for each unique class
for class_name in class_names :
    sheet = formattedWorkbook.create_sheet(title=class_name)
    #adding the headers to the worksheet
    sheet["A1"] = "Last Name"
    sheet["B1"] = "First Name"
    sheet["C1"] = "Student ID"
    sheet["D1"] = "Grade"
    #adding the class name to the class name list
    classNamesList.append(class_name)


#adding data to each sheet
#going through each class in the class list
for sClass in classNamesList :
    #setting the class sheet as
    sheet = formattedWorkbook[sClass]
    #going through each row in the bad data set
    for row in currSheet.iter_rows(min_row = 2, min_col = 1, max_row = currSheet.max_row, max_col = 3) :
        #checking to see if the class matches
        if (row[0].value == sClass) :
            #creating and infor list to store the data
            infoList = row[1].value.split("_")
            infoList.append(row[2].value)

            #adding the data to a new row in the class worksheet
            sheet.append(infoList)

# loop through each sheet 
for sheet_name in formattedWorkbook.sheetnames:
    currSheet = formattedWorkbook[sheet_name]
    # add a filter to each sheet
    currSheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(currSheet.max_column)}{currSheet.max_row}"

# Adding summary information to each class sheet
for sClass in classNamesList:
    sheet = formattedWorkbook[sClass]
    
    # Collect all grades (column D) as integers
    grades = [row[0] for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) if isinstance(row[0], (int, float))]

    # Organize header cells and columns into tuples to iterate
    tHeaders = ('A1', 'B1', 'C1', 'D1', 'F1', 'G1')
    tColumns = ('A', 'B', 'C', 'D', 'F', 'G')

    if grades:
        highest_grade = max(grades)
        lowest_grade = min(grades)
        mean_grade = sum(grades) / len(grades)
        median_grade = statistics.median(grades)
        student_count = len(grades)

        # Writing the summary information
        sheet["F1"] = "Summary"
        sheet["F2"] = "Highest Grade"
        sheet["G2"] = highest_grade
        sheet["F3"] = "Lowest Grade"
        sheet["G3"] = lowest_grade
        sheet["F4"] = "Mean Grade"
        sheet["G4"] = mean_grade
        sheet["F5"] = "Median Grade"
        sheet["G5"] = median_grade
        sheet["F6"] = "Number of Students"
        sheet["G6"] = student_count

    # Set header fonts to bold and resize columns
    for sHeader, sColumn in zip(tHeaders, tColumns):
        sheet[sHeader].font = Font(bold= True)

        if sheet[sHeader].value:
            sheet.column_dimensions[sColumn].width = len(sheet[sHeader].value) + 5
        else:
            sheet.column_dimensions[sColumn].width = 10
        
# Save the workbook and close it
formattedWorkbook.save(filename = 'formatted_grades.xlsx')
formattedWorkbook.close()